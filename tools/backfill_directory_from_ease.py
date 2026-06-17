"""One-time backfill of the directory from an EASE Company Directory export.

EASE is Theragen's HR system of record -- it has the real Department, Job Title,
and Location that Entra lacks. This matches EASE employees to directory persons
BY FULL NAME (EASE 'Email' is usually personal, not the work UPN), crosswalks the
granular EASE department to one of the 8 Theragen departments, and writes
Department + Location into the Staff Directory List (read back by sync_directory)
plus fills any blank person.job_title.

    python tools/backfill_directory_from_ease.py --file "<EASE.xlsx>" --dry-run
    python tools/backfill_directory_from_ease.py --file "<EASE.xlsx>"

Crosswalk note: nobody in EASE is in Clinical / R&D / IT / HR -- those 8-dept
buckets stay empty (the 8 are seed assumptions; EASE is a sales/fulfillment org).
"""
import argparse
import os
import re
import sys

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, os.path.join(ROOT, "tools"))

# EASE department (granular, numeric-prefixed) -> one of the 8 Theragen departments.
EASE_DEPARTMENT_CROSSWALK = {
    "100-Direct Sales": "Commercial / Marketing",
    "101-Sales Managers": "Commercial / Marketing",
    "101-Area Sales Director": "Commercial / Marketing",
    "102-Sales Trainers": "Commercial / Marketing",
    "8-Marketing": "Commercial / Marketing",
    "200-Order Fulfillment": "Operations / PMO",
    "202-Customer Care": "Operations / PMO",
    "6-Operations and Manufacturing": "Operations / PMO",
    "1-General Management": "Operations / PMO",
    "4-Insurance Contracting": "Operations / PMO",
    "201-Accounts Receivable": "Finance / Procurement",
    "7-Finance and Accounting": "Finance / Procurement",
    "9-Quality Control": "Regulatory / Quality",
    "9-Quality Specialist": "Regulatory / Quality",
}


def norm_name(first, last):
    return re.sub(r"\s+", " ", f"{first or ''} {last or ''}".strip().lower())


def crosswalk_department(ease_dept, crosswalk=EASE_DEPARTMENT_CROSSWALK):
    """EASE department string -> one of the 8 Theragen departments, or None (unmapped)."""
    return crosswalk.get((ease_dept or "").strip())


def read_ease(path):
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    # prefer a directory-looking sheet (a workbook may carry many sheets)
    name = next((s for s in wb.sheetnames if s.lower() in ("ease directory", "directory")),
                wb.sheetnames[0])
    rows = list(wb[name].iter_rows(values_only=True))
    idx = {h: i for i, h in enumerate(rows[0])}

    def cell(r, *keys):
        for k in keys:
            if k in idx and r[idx[k]] is not None:
                return str(r[idx[k]]).strip()
        return ""
    out = []
    for r in rows[1:]:
        first, last = cell(r, "First Name"), cell(r, "Last Name")
        if first or last:
            out.append({"name": norm_name(first, last), "display": f"{first} {last}".strip(),
                        "upn": cell(r, "User principal name", "User Principal Name", "UPN").lower(),
                        "title": cell(r, "Job Title"), "loc": cell(r, "Location"),
                        "dept": cell(r, "Department")})
    return out


def main(argv):
    ap = argparse.ArgumentParser()
    ap.add_argument("--file", required=True)
    ap.add_argument("--dry-run", action="store_true")
    args = ap.parse_args(argv)
    dry = args.dry_run
    import json
    import psycopg
    from graph_client import Graph
    import sync_directory as sd

    ease = read_ease(args.file)
    print(f"EASE employees: {len(ease)}")

    PG = json.load(open(os.path.join(ROOT, "db", ".pg.local.json"), encoding="utf-8"))
    M365 = json.load(open(os.path.join(ROOT, "db", ".m365.local.json"), encoding="utf-8"))
    with psycopg.connect(host=PG["server"], dbname=PG["database"], user=PG["user"],
                         password=PG["password"], sslmode="require", autocommit=True,
                         connect_timeout=20) as conn:
        people_by_name, people_by_upn = {}, {}
        for r in conn.execute("SELECT display_name, upn, email, job_title, person_id"
                              " FROM bi.org_directory").fetchall():
            rec = {"upn": r[1], "email": r[2], "title": r[3], "pid": r[4]}
            people_by_name[norm_name(r[0], "")] = rec
            if r[1]:
                people_by_upn[r[1].lower()] = rec
            if r[2]:
                people_by_upn[r[2].lower()] = rec

        g = Graph()
        site, lid = M365["site_id"], M365["staff_directory_list_id"]
        items = sd._list_items(g, site, lid, "UPN", "https://graph.microsoft.com/v1.0")
        item_by_upn = {((it.get("fields") or {}).get("UPN") or "").strip().lower(): it["id"] for it in items}

        matched = unmatched = no_item = unmapped = titles = 0
        dept_counts = {}
        for e in ease:
            dp = (people_by_upn.get(e["upn"]) if e.get("upn") else None) or people_by_name.get(e["name"])
            if not dp:
                unmatched += 1
                continue
            matched += 1
            upn = (dp["upn"] or dp["email"] or "").strip().lower()
            item_id = item_by_upn.get(upn)
            dept8 = crosswalk_department(e["dept"])
            if dept8:
                dept_counts[dept8] = dept_counts.get(dept8, 0) + 1
            else:
                unmapped += 1
            fields = {}
            if dept8 and item_id:
                fields["Department"] = dept8
            if e["loc"] and item_id:
                fields["Location"] = e["loc"]
            if not item_id:
                no_item += 1
            elif fields and not dry:
                g.patch(f"/sites/{site}/lists/{lid}/items/{item_id}/fields", fields)
            if e["title"] and not dp["title"]:
                titles += 1
                if not dry:
                    conn.execute("UPDATE doc_mgmt.person SET job_title=%s WHERE person_id=%s",
                                 (e["title"], dp["pid"]))

    print(f"matched by name: {matched} | unmatched: {unmatched} | no list item: {no_item}")
    print(f"department crosswalked: {sum(dept_counts.values())} | unmapped EASE dept: {unmapped}")
    print(f"job titles filled (were blank): {titles}")
    print("--- department distribution (crosswalked) ---")
    for d, c in sorted(dept_counts.items(), key=lambda x: -x[1]):
        print(f"  {c:>2}  {d}")
    print("DRY-RUN - no writes." if dry else "done. (run sync_directory to read Department/Location back)")
    return 0


if __name__ == "__main__":
    sys.exit(main(sys.argv[1:]))
