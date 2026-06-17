"""Import the field-sales distributor network from the 'Master DIst' sheet.

The source (an Excel workbook maintained by sales ops) is messy: free-text Status,
RSM typos, blank regions, ~606 rows for ~329 distinct distributors. This reads the
Master DIst sheet, normalizes Status/RSM/Region, dedups to one row per distributor
(keeping the most-complete row), and upserts sales.distributor (read back by
bi.distributor). Pure normalizers are unit-tested; heavy imports are lazy.

    python tools/import_distributors.py --file "<workbook.xlsx>" --dry-run
    python tools/import_distributors.py --file "<workbook.xlsx>"
"""
import argparse
import os
import sys
import uuid

NS = uuid.NAMESPACE_URL

# RSM typo / alias fixes -> canonical name (keyed lower-cased).
RSM_CANONICAL = {"tom milory": "Tom Milroy"}


def normalize_distributor_status(raw):
    """Free-text Status -> Active / Terminated / Inactive / Pending / Unknown."""
    s = (raw or "").strip().lower()
    if not s:
        return "Unknown"
    if "termin" in s or s == "gone":
        return "Terminated"
    if "do not pay" in s:
        return "Inactive"
    if "pending" in s:
        return "Pending"
    return "Active"


def canonical_rsm(raw):
    s = (raw or "").strip()
    if not s:
        return "Unknown"
    return RSM_CANONICAL.get(s.lower(), s)


def normalize_region(raw):
    s = (raw or "").strip()
    return s if s else "Unknown"


def dedupe_distributors(rows):
    """One row per distinct distributor name (lower-cased); keep the most-complete
    row (most non-blank region/rsm/status/products/email). Drops blank-name rows."""
    def score(r):
        return sum(1 for k in ("region", "rsm", "status_raw", "products", "email") if r.get(k))
    best = {}
    for r in rows:
        key = (r.get("name") or "").strip().lower()
        if not key:
            continue
        if key not in best or score(r) > score(best[key]):
            best[key] = r
    return list(best.values())


def read_master_dist(path):
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    name = next((s for s in wb.sheetnames if "master" in s.lower() and "dist" in s.lower()),
                wb.sheetnames[0])
    rows = list(wb[name].iter_rows(values_only=True))
    idx = {h: i for i, h in enumerate(rows[0])}

    def cell(r, *keys):
        for k in keys:
            if k in idx and r[idx[k]] is not None:
                return str(r[idx[k]]).strip()
        return ""
    out = []
    for r in rows[1:]:
        if cell(r, "Distributor Name"):
            out.append({
                "name": cell(r, "Distributor Name"),
                "marketing_rep": cell(r, "Marketing Rep Full Name"),
                "status_raw": cell(r, "Status"),
                "sales_type": cell(r, "Sales Type"),
                "region": cell(r, "Region"),
                "rsm": cell(r, "RSM"),
                "products": cell(r, "Products"),
                "email": cell(r, "Email"),
            })
    return out


def main(argv):
    ap = argparse.ArgumentParser()
    ap.add_argument("--file", required=True)
    ap.add_argument("--dry-run", action="store_true")
    args = ap.parse_args(argv)
    dry = args.dry_run
    import json
    import psycopg
    from collections import Counter

    raw = read_master_dist(args.file)
    rows = dedupe_distributors(raw)
    print(f"Master DIst rows: {len(raw)} -> {len(rows)} distinct distributors")

    PG = json.load(open(os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
                                     "db", ".pg.local.json"), encoding="utf-8"))
    status_c, region_c, active_c = Counter(), Counter(), 0
    with psycopg.connect(host=PG["server"], dbname=PG["database"], user=PG["user"],
                         password=PG["password"], sslmode="require", autocommit=True,
                         connect_timeout=20) as conn:
        for r in rows:
            ext = r["name"].strip().lower()
            status = normalize_distributor_status(r["status_raw"])
            region = normalize_region(r["region"])
            rsm = canonical_rsm(r["rsm"])
            active = status == "Active"
            status_c[status] += 1
            region_c[region] += 1
            active_c += 1 if active else 0
            if not dry:
                conn.execute(
                    "INSERT INTO sales.distributor (distributor_id, name, sales_type, region,"
                    " rsm, marketing_rep, status_raw, status, active, products, email, external_ref)"
                    " VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)"
                    " ON CONFLICT (external_ref) DO UPDATE SET name=EXCLUDED.name,"
                    " sales_type=EXCLUDED.sales_type, region=EXCLUDED.region, rsm=EXCLUDED.rsm,"
                    " marketing_rep=EXCLUDED.marketing_rep, status_raw=EXCLUDED.status_raw,"
                    " status=EXCLUDED.status, active=EXCLUDED.active, products=EXCLUDED.products,"
                    " email=EXCLUDED.email",
                    (str(uuid.uuid5(NS, "thg/distributor/" + ext)), r["name"], r["sales_type"] or None,
                     region, rsm, r["marketing_rep"] or None, r["status_raw"] or None, status, active,
                     r["products"] or None, r["email"] or None, ext))
    print(f"active: {active_c} | by status: {dict(status_c.most_common())}")
    print(f"by region: {dict(region_c.most_common())}")
    print("DRY-RUN - no writes." if dry else "done.")
    return 0


if __name__ == "__main__":
    sys.exit(main(sys.argv[1:]))
